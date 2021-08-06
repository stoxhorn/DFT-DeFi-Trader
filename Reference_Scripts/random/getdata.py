import os
from binance.client import Client
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as get_cl
from datetime import datetime
from openpyxl.styles import Color, PatternFill, Font, Border

filename = "C:\\Users\\Stoxhorn\\Desktop\\Vigtige_filer\\crypto\\dt.xlsx"

red = Color(rgb='00FF0000')
red_fill = PatternFill(patternType='solid', fgColor=red)

green = Color(rgb='000CCA3E')
green_fill = PatternFill(patternType='solid', fgColor=green)

black = Color(rgb='00000000')
black_fill = PatternFill(patternType='solid', fgColor=green)

def get_free(asset):
    return float(asset["free"])

def get_locked(asset):
    return float(asset["locked"])

def get_holding(ticker, amount, client):
    price = -1
    if ticker == "USDT":
        price = amount
    else:
        try:
            respone = client.get_symbol_ticker(symbol=(ticker + "USDT"))
            price = float(respone["price"])
        except:
            try:
                response = client.get_symbol_ticker(symbol=(ticker + "BTC"))
                btc_price = client.get_symbol_ticker(symbol=("BTC" + "USDT"))
                value = float(btc_price["price"]) * float(response["price"])
                price =  value

            except:
                try:
                    response = client.get_symbol_ticker(symbol=("USDT" + ticker))
                    price =  1/(float(response["price"]))

                except:
                    print("missing proper combo for: " + ticker)
    return [ticker, amount, price, amount * price]

def get_holdings():
    f = open("api.txt","r")
    # init
    api_key = f.readline()[:-1]
    # if no empty line afterwards, remove -1
    api_secret = f.readline()[:-1]

    client = Client(api_key, api_secret)
    f.close()
    bal = client.get_account()
    assets = []
    for i in bal["balances"]:
        if float(i["free"]) > 0.:
            assets.append(i)
        elif float(i["locked"]) > 0.:
            assets.append(i)
        else:
            pass

    holding = {}

    for i in assets:
        if get_free(i) > 0.:
            ticker = i["asset"]
            amount = get_free(i)
            price = get_holding(ticker, amount, client)
            holding[price[0]] = price
        elif get_locked(i) > 0.:
            ticker = i["asset"]
            amount = get_locked(i)
            price = get_holding(ticker, amount, client)
            holding[price[0]] = price
    return holding
    
def update_trade():


    hold = get_holdings()

    workbook = load_workbook(filename=filename)
    sheet = workbook["CoinHODL"]

    end = sheet["A1"].value.split("-")
    endc = end[0]
    endr = int(end[1])

    endc_num = sheet[endc + str(endr)].col_idx

    cur_col = endc_num +1
    cur_row = 4

    sheet["A1"] = get_cl(endc_num+3) + "-" + str(endr)
    while cur_row < (endr+1):
        sheet[get_cl(endr+3) + str(cur_row)].fill = black_fill
        cur_row += 1
    
    cur_row = 4

    for asset in hold:
        for i in range(endr):
            i += 4
            # if not in the row corresponding to the coin
            if hold[asset][0] != sheet["B" + str(i)].value:
                pass
            else:
                # if in the right row
                sheet[get_cl(cur_col) + str(i)] = hold[asset][2]
                # calculate percentage increase from original price
                sheet[get_cl(cur_col+1) + str(i)] = ((hold[asset][2]/float(sheet[get_cl(endc_num - 2) + str(i)].value))-1)*100

                cur_cell = get_cl(cur_col) + str(i)
                add_prof_perc(i, sheet, cur_cell, get_cl(cur_col+1) + str(i))
                add_prof(i, sheet, hold, asset, cur_cell, get_cl(cur_col+1) + str(i))
                add_worth(sheet, i, hold, asset, cur_cell, get_cl(cur_col+1) + str(i))
    


    avg_list = ["C", "D", "F", get_cl(endc_num+1)]
    sum_list = ["E", "G", "H"]

    cur_row = 4
    sum = 0
    sumold = 0
    test = []
    while cur_row < (endr+1):
        ratio = (((sheet[get_cl(cur_col+1) + str(cur_row)].value)/100))
        sum += ratio * float(sheet["E" + str(cur_row)].value)
        sumold += float(sheet["E" + str(cur_row)].value)
        cur_row += 1

    perc_prof = (sum/sumold)*100
    sheet[get_cl(cur_col+1) + str(cur_row)] = perc_prof
    if perc_prof > 0:
        sheet[get_cl(endr+3) + str(cur_row)].fill = green_fill
        sheet[get_cl(endr+3) + str(cur_row)].font = Font(color = "000000")
        pass


    for i in avg_list:
        write_avg_col(i, endr, sheet)
    for i in sum_list:
        write_sum_col(i, endr, sheet)
    time = datetime.now()
    sheet[get_cl(cur_col) + str(3)] = str(time)


    workbook.save(filename=filename)



def add_worth(sheet, i, hold, asset, cur_cell, next_cell):
    sheet["H" + str(i)] = hold[asset][3] #str(hold[asset][3]).replace(".", ",")
    if float(sheet[cur_cell].value) > float(sheet["C" + str(i)].value):
        sheet["H" + str(i)].fill = green_fill
        sheet[cur_cell].fill = green_fill
        sheet[cur_cell].font = Font(color = "000000")
        sheet[next_cell].fill = green_fill
        sheet[next_cell].font = Font(color = "000000")
    else:
        sheet["H" + str(i)].fill = red_fill
        sheet[cur_cell].fill = red_fill
        sheet[cur_cell].font = Font(color = "000000")
        sheet[next_cell].fill = red_fill
        sheet[next_cell].font = Font(color = "000000")
    

def add_prof_perc(i, sheet, cur_cell, next_cell):
    sheet["F" + str(i)] = "=((" + cur_cell + "/C" + str(i) + ") -1) * 100"
    if float(sheet[cur_cell].value) > float(sheet["C" + str(i)].value):
        sheet["F" + str(i)].fill = green_fill
        sheet[cur_cell].fill = green_fill
        sheet[cur_cell].font = Font(color = "000000")
        sheet[next_cell].fill = green_fill
        sheet[next_cell].font = Font(color = "000000")
    else:
        sheet["F" + str(i)].fill = red_fill
        sheet[cur_cell].fill = red_fill
        sheet[cur_cell].font = Font(color = "000000")
        sheet[next_cell].fill = red_fill
        sheet[next_cell].font = Font(color = "000000")

def add_prof(i, sheet, hold, asset, cur_cell, next_cell): 
    sheet["G" + str(i)] = "=" + str(hold[asset][3]) + "-E" + str(i)
    if float(sheet[cur_cell].value) > float(sheet["C" + str(i)].value):
        sheet["G" + str(i)].fill = green_fill
        sheet[cur_cell].fill = green_fill
        sheet[cur_cell].font = Font(color = "000000")
        sheet[next_cell].fill = green_fill
        sheet[next_cell].font = Font(color = "000000")
    else:
        sheet["G" + str(i)].fill = red_fill
        sheet[cur_cell].fill = red_fill
        sheet[cur_cell].font = Font(color = "000000")
        sheet[next_cell].fill = red_fill
        sheet[next_cell].font = Font(color = "000000")

        
def write_avg_col(col, endr, sheet):
    sheet[col + str(endr+1)] = "=SUM(" + col + "4:" + col + str(endr) + ")/" + str(endr-3)

def write_sum_col(col, endr, sheet):
    sheet[col + str(endr+1)] = "=SUM(" + col + "4:" + col + str(endr) + ")"

if __name__ == "__main__":
    update_trade()







